from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.db.models import Q, F
import json
import requests
import openpyxl
from openpyxl import Workbook
from datetime import datetime, timedelta
from .models import Item, WebhookSettings, ActivityLog
from .forms import LoginForm, ItemForm, WebhookSettingsForm, CustomPasswordChangeForm, ExcelUploadForm

def login_view(request):
    if request.user.is_authenticated:
        return redirect('inventory:dashboard')
    
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            user = authenticate(request, username=username, password=password)
            
            if user is not None:
                login(request, user)
                ActivityLog.objects.create(
                    user=user,
                    action='Login',
                    status='success',
                    notes='Login berhasil'
                )
                return redirect('inventory:dashboard')
            else:
                messages.error(request, 'Username atau password salah')
                ActivityLog.objects.create(
                    user=None,
                    action='Login',
                    status='failed',
                    notes=f'Percobaan login gagal untuk username: {username}'
                )
    else:
        form = LoginForm()
    
    return render(request, 'inventory/login.html', {'form': form})

@login_required
def logout_view(request):
    ActivityLog.objects.create(
        user=request.user,
        action='Logout',
        status='success',
        notes='Logout berhasil'
    )
    logout(request)
    return redirect('inventory:login')

@login_required
def dashboard(request):
    query = request.GET.get('query', '')
    sort = request.GET.get('sort', '')
    sort_dir = request.GET.get('dir', 'asc')
    filter_type = request.GET.get('filter', '')
    
    # Start with all items or filtered by search query
    if query:
        items = Item.objects.filter(
            Q(code__icontains=query) | 
            Q(name__icontains=query) | 
            Q(category__icontains=query)
        )
    else:
        items = Item.objects.all()
    
    # Apply filters
    if filter_type == 'low_stock':
        # Filter items with stock below minimum
        items = items.filter(current_stock__lt=F('minimum_stock'))
    
    # Apply sorting
    if sort:
        # Determine sort field and direction
        sort_field = ''
        if sort == 'name':
            sort_field = 'name'
        elif sort == 'name_desc':
            sort_field = '-name'
        elif sort == 'category':
            sort_field = 'category'
        elif sort == 'category_desc':
            sort_field = '-category'
        elif sort == 'stock_asc':
            sort_field = 'current_stock'
        elif sort == 'stock_desc':
            sort_field = '-current_stock'
        elif sort == 'price_asc':
            sort_field = 'selling_price'
        elif sort == 'price_desc':
            sort_field = '-selling_price'
        
        # Apply sorting if a valid field was specified
        if sort_field:
            items = items.order_by(sort_field)
    
    return render(request, 'inventory/dashboard.html', {
        'items': items,
        'query': query,
        'current_sort': sort,
        'current_filter': filter_type,
    })

@login_required
def upload_file(request):
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            
            try:
                wb = openpyxl.load_workbook(excel_file)
                sheet = wb.active
                
                # Skip header row
                rows = list(sheet.rows)[1:]
                
                success_count = 0
                error_count = 0
                
                for row in rows:
                    try:
                        code = str(row[0].value).strip()
                        name = str(row[1].value).strip()
                        category = str(row[2].value).strip()
                        # Swap columns 3 and 4 to fix the issue
                        selling_price = float(row[3].value) if row[3].value is not None else 0
                        current_stock = int(row[4].value) if row[4].value is not None else 0
                        
                        # Check if item exists
                        try:
                            item = Item.objects.get(code=code)
                            # Update existing item but preserve minimum_stock
                            item.name = name
                            item.category = category
                            item.current_stock = current_stock
                            item.selling_price = selling_price
                            item.save()
                        except Item.DoesNotExist:
                            # Create new item
                            Item.objects.create(
                                code=code,
                                name=name,
                                category=category,
                                current_stock=current_stock,
                                selling_price=selling_price,
                                minimum_stock=0  # Default value
                            )
                        
                        success_count += 1
                    except Exception as e:
                        error_count += 1
                        print(f"Error processing row: {e}")
                
                ActivityLog.objects.create(
                    user=request.user,
                    action='Upload File',
                    status='success',
                    notes=f'Berhasil memproses {success_count} item, gagal {error_count} item'
                )
                
                messages.success(request, f'File berhasil diupload. {success_count} item diproses, {error_count} item gagal.')
                return redirect('inventory:dashboard')
            
            except Exception as e:
                ActivityLog.objects.create(
                    user=request.user,
                    action='Upload File',
                    status='failed',
                    notes=f'Gagal memproses file: {str(e)}'
                )
                messages.error(request, f'Gagal memproses file: {str(e)}')
    else:
        form = ExcelUploadForm()
    
    return render(request, 'inventory/upload_file.html', {'form': form})

@login_required
def backup_file(request):
    items = Item.objects.all()
    
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"
    
    # Add headers
    headers = ['Kode Barang', 'Nama Barang', 'Kategori', 'Stok Saat Ini', 'Harga Jual', 'Stok Minimum']
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
    
    # Add data
    for row_num, item in enumerate(items, 2):
        ws.cell(row=row_num, column=1, value=item.code)
        ws.cell(row=row_num, column=2, value=item.name)
        ws.cell(row=row_num, column=3, value=item.category)
        ws.cell(row=row_num, column=4, value=item.current_stock)
        ws.cell(row=row_num, column=5, value=float(item.selling_price))
        ws.cell(row=row_num, column=6, value=item.minimum_stock)
    
    # Create response
    timestamp = datetime.now().strftime('%Y_%m_%d_%H_%M_%S')
    filename = f"Backup_Inventory_{timestamp}.xlsx"
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save the workbook to the response
    wb.save(response)
    
    ActivityLog.objects.create(
        user=request.user,
        action='Backup File',
        status='success',
        notes=f'Backup file berhasil dibuat: {filename}'
    )
    
    return response

@login_required
def change_password(request):
    if request.method == 'POST':
        form = CustomPasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            
            ActivityLog.objects.create(
                user=request.user,
                action='Ubah Password',
                status='success',
                notes='Password berhasil diubah'
            )
            
            messages.success(request, 'Password berhasil diubah!')
            return redirect('inventory:dashboard')
        else:
            ActivityLog.objects.create(
                user=request.user,
                action='Ubah Password',
                status='failed',
                notes='Gagal mengubah password'
            )
    else:
        form = CustomPasswordChangeForm(request.user)
    
    return render(request, 'inventory/change_password.html', {'form': form})

@login_required
def webhook_settings(request):
    try:
        webhook_settings = WebhookSettings.objects.first()
        if not webhook_settings:
            webhook_settings = WebhookSettings.objects.create()
    except:
        webhook_settings = WebhookSettings.objects.create()
    
    if request.method == 'POST':
        form = WebhookSettingsForm(request.POST, instance=webhook_settings)
        if form.is_valid():
            webhook = form.save(commit=False)
            webhook.updated_by = request.user
            webhook.save()
            
            ActivityLog.objects.create(
                user=request.user,
                action='Update Webhook',
                status='success',
                notes=f'Webhook URL diperbarui: {webhook.telegram_webhook_url}'
            )
            
            messages.success(request, 'Webhook settings berhasil diperbarui!')
            return redirect('inventory:webhook_settings')
    else:
        form = WebhookSettingsForm(instance=webhook_settings)
    
    return render(request, 'inventory/webhook_settings.html', {'form': form})

@login_required
def activity_logs(request):
    # Get logs from the last 7 days
    seven_days_ago = timezone.now() - timedelta(days=7)
    logs = ActivityLog.objects.filter(timestamp__gte=seven_days_ago).order_by('-timestamp')
    
    return render(request, 'inventory/activity_logs.html', {'logs': logs})

@login_required
def send_to_telegram(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            item_ids = data.get('item_ids', [])
            
            if not item_ids:
                return JsonResponse({'status': 'error', 'message': 'Tidak ada item yang dipilih'}, status=400)
            
            items = Item.objects.filter(id__in=item_ids)
            
            # Get webhook URL
            webhook_settings = WebhookSettings.objects.first()
            if not webhook_settings or not webhook_settings.telegram_webhook_url:
                return JsonResponse({'status': 'error', 'message': 'URL Webhook belum diatur'}, status=400)
            
            # Prepare data for Zapier - Simplified format
            products_list = []
            for item in items:
                products_list.append({
                    'kode': item.code,
                    'nama': item.name,
                    'kategori': item.category,
                    'stok': item.current_stock,
                    'harga': int(item.selling_price),
                    'stok_minimum': item.minimum_stock
                })
            
            # Try multiple payload formats to ensure compatibility with Zapier
            
            # Format 1: Simple array of products
            payload1 = products_list
            
            # Format 2: Object with products array
            payload2 = {
                'products': products_list
            }
            
            # Format 3: Single product (if only one selected)
            payload3 = products_list[0] if products_list else {}
            
            # Log all payloads for debugging
            print(f"Webhook URL: {webhook_settings.telegram_webhook_url}")
            print(f"Payload 1 (array): {json.dumps(payload1)}")
            print(f"Payload 2 (object): {json.dumps(payload2)}")
            print(f"Payload 3 (single): {json.dumps(payload3)}")
            
            # Common headers
            headers = {
                'Content-Type': 'application/json',
                'User-Agent': 'Stock Management System/1.0'
            }
            
            # Try all payload formats
            success = False
            response_logs = []
            
            # Try Format 1
            try:
                response1 = requests.post(
                    webhook_settings.telegram_webhook_url,
                    json=payload1,
                    headers=headers,
                    timeout=10
                )
                response_logs.append(f"Format 1 response: {response1.status_code} - {response1.text}")
                if response1.status_code in [200, 201, 202]:
                    success = True
            except Exception as e:
                response_logs.append(f"Format 1 error: {str(e)}")
            
            # Try Format 2 if not successful yet
            if not success:
                try:
                    response2 = requests.post(
                        webhook_settings.telegram_webhook_url,
                        json=payload2,
                        headers=headers,
                        timeout=10
                    )
                    response_logs.append(f"Format 2 response: {response2.status_code} - {response2.text}")
                    if response2.status_code in [200, 201, 202]:
                        success = True
                except Exception as e:
                    response_logs.append(f"Format 2 error: {str(e)}")
            
            # Try Format 3 if not successful yet and we have only one product
            if not success and len(products_list) == 1:
                try:
                    response3 = requests.post(
                        webhook_settings.telegram_webhook_url,
                        json=payload3,
                        headers=headers,
                        timeout=10
                    )
                    response_logs.append(f"Format 3 response: {response3.status_code} - {response3.text}")
                    if response3.status_code in [200, 201, 202]:
                        success = True
                except Exception as e:
                    response_logs.append(f"Format 3 error: {str(e)}")
            
            # Try with form-encoded data as last resort
            if not success:
                try:
                    form_data = {'payload': json.dumps(payload2)}
                    response4 = requests.post(
                        webhook_settings.telegram_webhook_url,
                        data=form_data,
                        timeout=10
                    )
                    response_logs.append(f"Form data response: {response4.status_code} - {response4.text}")
                    if response4.status_code in [200, 201, 202]:
                        success = True
                except Exception as e:
                    response_logs.append(f"Form data error: {str(e)}")
            
            # Log all responses
            for log in response_logs:
                print(log)
            
            if success:
                # Create a list of product names for the log
                product_names = [item.name for item in items]
                product_names_str = ", ".join(product_names)
                
                ActivityLog.objects.create(
                    user=request.user,
                    action='Kirim ke Telegram',
                    status='success',
                    notes=f'Berhasil mengirim {len(items)} item ke Telegram: {product_names_str}'
                )
                return JsonResponse({'status': 'success', 'message': 'Data berhasil dikirim ke Telegram'})
            else:
                # Create a list of product names for the error log
                product_names = [item.name for item in items]
                product_names_str = ", ".join(product_names)
                
                ActivityLog.objects.create(
                    user=request.user,
                    action='Kirim ke Telegram',
                    status='failed',
                    notes=f'Gagal mengirim item ke Telegram: {product_names_str}. Logs: {"; ".join(response_logs)}'
                )
                return JsonResponse({'status': 'error', 'message': 'Gagal mengirim data. Lihat log untuk detail.'}, status=400)
                
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            print(f"Error sending to Telegram: {error_details}")
            
            # Try to get product names if possible
            product_names_str = "Unknown products"
            try:
                if 'item_ids' in data:
                    items = Item.objects.filter(id__in=data['item_ids'])
                    product_names = [item.name for item in items]
                    product_names_str = ", ".join(product_names)
            except:
                pass
            
            ActivityLog.objects.create(
                user=request.user,
                action='Kirim ke Telegram',
                status='failed',
                notes=f'Error saat mengirim item: {product_names_str}. Detail: {str(e)}'
            )
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    
    return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)

@login_required
def update_min_stock(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            item_id = data.get('item_id')
            minimum_stock = data.get('minimum_stock', 0)
            
            if not item_id:
                return JsonResponse({'status': 'error', 'message': 'ID item tidak ditemukan'}, status=400)
            
            try:
                item = Item.objects.get(id=item_id)
                item.minimum_stock = minimum_stock
                item.save()
                
                ActivityLog.objects.create(
                    user=request.user,
                    action='Update Stok Minimum',
                    status='success',
                    notes=f'Stok minimum untuk {item.name} diubah menjadi {minimum_stock}'
                )
                
                return JsonResponse({'status': 'success', 'message': 'Stok minimum berhasil diperbarui'})
            except Item.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'Item tidak ditemukan'}, status=404)
                
        except Exception as e:
            ActivityLog.objects.create(
                user=request.user,
                action='Update Stok Minimum',
                status='failed',
                notes=f'Error: {str(e)}'
            )
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    
    return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)

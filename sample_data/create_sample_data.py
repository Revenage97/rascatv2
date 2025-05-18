from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import os

# Create a new workbook
wb = Workbook()
ws = wb.active
ws.title = "Inventory"

# Add headers with styling
headers = ['Kode', 'Nama Barang', 'Kategori', 'Total Stok', 'Harga Jual']
header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
header_font = Font(bold=True)

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')

# Add sample data
sample_data = [
    ['0000000000001', 'Obat Kutu Detick Kucing Anjing 3ml', 'detick', 45, 36900],
    ['0000000000002', 'Vitamin Kucing Growth Booster 60ml', 'vitamin', 30, 42500],
    ['0000000000003', 'Makanan Kucing Royal Canin 1kg', 'makanan', 15, 89000],
    ['0000000000004', 'Pasir Kucing Gumpal Wangi 10L', 'pasir', 20, 55000],
    ['0000000000005', 'Sisir Kutu Kucing dan Anjing', 'grooming', 12, 15000],
    ['0000000000006', 'Shampoo Kucing Anti Kutu 200ml', 'grooming', 18, 38500],
    ['0000000000007', 'Mainan Kucing Bola dengan Lonceng', 'mainan', 25, 12000],
    ['0000000000008', 'Kandang Kucing Lipat Portable', 'kandang', 8, 175000],
    ['0000000000009', 'Obat Tetes Mata Kucing 10ml', 'obat', 14, 45000],
    ['0000000000010', 'Susu Kucing Anakan 250gr', 'makanan', 22, 65000],
]

for row_num, row_data in enumerate(sample_data, 2):
    for col_num, cell_value in enumerate(row_data, 1):
        ws.cell(row=row_num, column=col_num, value=cell_value)

# Adjust column widths
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        if cell.value:
            cell_length = len(str(cell.value))
            if cell_length > max_length:
                max_length = cell_length
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

# Save the workbook
file_path = os.path.join(os.path.dirname(__file__), 'sample_inventory.xlsx')
wb.save(file_path)

print(f"Sample inventory Excel file created at {file_path}")
